#!/usr/bin/env python3
"""Construit data/seed.json à partir du fichier réel BECLOUD ANALYSE.xlsx.

Fusionne les onglets « Liste Users » (identité, taille boîte, cases licence) et
« Suivi Licences » (profil P1..P4b, engagement, paiement, migration documentaire).
Produit l'état initial complet de l'application (entités + paramètres + jalons).
"""
import json, sys, datetime, re, os

import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else \
    "/root/.claude/uploads/01bde471-b41d-5b99-9f1d-a903c3f7cd1e/ab5b4094-BECLOUD_ANALYSE.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else \
    os.path.join(os.path.dirname(__file__), "..", "data", "seed.json")

wb = openpyxl.load_workbook(SRC, data_only=True)


def iso(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return None


# --- Onglet Liste Users : identité + taille + cases licence -----------------
lu = list(wb["Liste Users"].iter_rows(values_only=True))
users_raw = []
for r in lu[1:46]:  # r1..r45 = 45 comptes
    if not r or not r[2]:
        continue
    users_raw.append({
        "first": (r[0] or "").strip(),
        "last": (r[1] or "").strip(),
        "email": (r[2] or "").strip(),
        "lastSignIn": iso(r[4]),
        "usage": round(float(r[5]), 2) if r[5] not in (None, "") else 0.0,
        "cb_2go": bool(r[6]),
        "cb_50go": bool(r[7]),
        "cb_2gof3": bool(r[8]),
        "cb_50gof3": bool(r[9]),
        "cb_business": bool(r[10]),
    })

# --- Onglet Suivi Licences : profil / engagement / paiement / doc migration -
sl = list(wb["Suivi Licences"].iter_rows(values_only=True))
suivi = {}
for r in sl[14:60]:
    if not r or not r[3]:
        continue
    email = (r[3] or "").strip().lower()
    profile = (r[6] or "").strip() if r[6] else ""
    suivi[email] = {
        "profile": profile,
        "takeout": (str(r[9]).strip().lower() == "oui") if r[9] else False,
        "nas": (str(r[10]).strip().lower() == "oui") if r[10] else False,
        "engagement": (str(r[11]).strip().lower() if r[11] else "annuel"),
        "payment": (str(r[12]).strip().lower() if r[12] else "mensuel"),
    }

# --- Inférences --------------------------------------------------------------
SERVICE_PATTERNS = re.compile(
    r"^(maison\.|contact|office|scan|eshop|e-shop|communication|coupe|"
    r"atelier\.|gm-|.*\.order$|corthay\.order|japan\.order|assistantretail)",
    re.I,
)


def infer_site(email, first, last):
    e = email.lower()
    n = f"{first} {last}".lower()
    if "london" in e or "london" in n or "motconb" in e or "gm-corthaylondon" in e:
        return "Londres"
    if any(k in e for k in ["aoyama", "osaka", "shinjuku", "japan"]) or "osaka" in n:
        return "Japon"
    if "korea" in e or "korea" in n or "cheongdam" in e:
        return "Corée"
    if "beijing" in e or "china" in e or "zhongya" in n:
        return "Chine"
    if "assistantretail" in e:
        return "Hong Kong"
    if "atelier" in e or "coupe" in e:
        return "Manufacture"
    return "Paris"


def infer_profile(email, u):
    s = suivi.get(email.lower())
    if s and s["profile"]:
        return s["profile"].replace(" ", "")
    if u["cb_business"]:
        return "P1"
    if u["cb_50gof3"]:
        return "P2"
    if u["cb_2gof3"]:
        return "P3"
    if u["cb_50go"]:
        return "P4a"
    if u["cb_2go"]:
        return "P4b"
    return "P4a"  # défaut raisonnable pour les comptes « à arbitrer »


MAIL_STATUSES = ["non commencé", "copie lancée", "copié", "basculé", "reconnecté", "validé"]
MFA_STATUSES = [
    ("non démarrée", "non défini", False),
    ("à faire", "non défini", False),
    ("configurée Authenticator", "Authenticator", True),
    ("configurée SMS", "SMS", True),
    ("bloquée", "Authenticator", False),
]
COMM_STATUSES = ["non démarré", "email envoyé", "relancé", "confirmé"]

users = []
for i, u in enumerate(users_raw):
    email = u["email"]
    local = email.split("@")[0]
    s = suivi.get(email.lower(), {})
    profile = infer_profile(email, u)
    is_service = bool(SERVICE_PATTERNS.match(local))
    is_arbitrer = email.lower() in suivi and not suivi[email.lower()]["profile"]

    name = f"{u['first']} {u['last']}".strip().lower()
    vip = "xavier" in email.lower() or "pierre" in email.lower()

    # Statut du compte
    if email.lower() in ("communication@corthay.com", "gm-corthaylondon@corthay.com",
                          "maison.beijing@corthay.com", "maison.china@corthay.com"):
        status = "ancien salarié"
        is_service = True
    elif is_service:
        status = "boîte de service"
    elif is_arbitrer:
        status = "à arbitrer"
    else:
        status = "actif"

    # Variété déterministe des statuts (selon index)
    mail = MAIL_STATUSES[i % len(MAIL_STATUSES)]
    mfa_status, mfa_method, mfa_conf = MFA_STATUSES[i % len(MFA_STATUSES)]
    comm = COMM_STATUSES[i % len(COMM_STATUSES)]

    physical = not is_service and status not in ("ancien salarié",)

    # Boîte
    scan = local.lower().startswith("scan")
    auto_send = local.lower() in ("eshop", "e-shop", "corthay.order", "japan.order", "contact")
    size = u["usage"]
    if profile == "SHARED" or (is_service and size <= 50 and not scan and not auto_send):
        type_target = "boîte partagée"
    elif status == "ancien salarié":
        type_target = "à archiver"
    elif scan or auto_send:
        type_target = "utilisateur Microsoft"
    else:
        type_target = "utilisateur Microsoft"

    type_current = (
        "scan-to-mail" if scan else
        "boutique" if local.startswith("maison.") else
        "boîte service" if is_service else
        "ancien salarié" if status == "ancien salarié" else
        "utilisateur Google"
    )

    # Nettoyage pour les grosses boîtes
    cleanup_req = size > 50
    cleanup_done = False
    target_size = 45.0 if size > 50 else None

    risk = "rouge" if size > 100 else ("orange" if size > 50 else "vert")
    if vip and mail != "validé":
        risk = "rouge" if size > 100 else "orange"

    users.append({
        "id": f"u{i+1:03d}",
        "firstName": u["first"],
        "lastName": u["last"],
        "googleEmail": email,
        "microsoftEmail": email,
        "status": status,
        "site": infer_site(email, u["first"], u["last"]),
        "role": "",
        "phone": "",
        "vip": vip,
        "physicalUser": physical,
        "usesOutlookWeb": physical,
        "usesOutlookDesktop": physical and (i % 2 == 0),
        "usesMobile": physical and (i % 3 == 0),
        "os": ["Windows", "Mac", "inconnu"][i % 3],
        "lastGoogleSignIn": u["lastSignIn"],
        "mailboxSizeGB": size,
        "targetSizeGB": target_size,
        "cleanupRequested": cleanup_req,
        "cleanupDone": cleanup_done,
        "mailStatus": "non commencé" if status == "ancien salarié" else mail,
        "mfa": {
            "status": "non démarrée" if not physical else mfa_status,
            "method": mfa_method,
            "configured": physical and mfa_conf,
            "configuredAt": None,
            "blocked": physical and mfa_status == "bloquée",
            "needsAssistance": physical and mfa_status in ("bloquée", "à faire"),
            "instructionSent": physical and comm != "non démarré",
            "firstSignInDone": mail not in ("non commencé", "copie lancée"),
        },
        "commStatus": comm,
        "remarks": "Migration documentaire (Takeout/NAS) à prévoir." if s.get("takeout") else "",
        "linkedMailboxes": [],
        "risk": risk,
        "licenseProfile": profile,
        "engagement": "annuel" if s.get("engagement", "annuel").startswith("ann") else "mensuel",
        "payment": "mensuel",
        "packBeCloud": profile == "P1",
        "mailbox": {
            "typeCurrent": type_current,
            "typeTarget": type_target,
            "members": [],
            "memberRight": "accès complet",
            "alias": [],
            "autoSend": auto_send,
            "scanToMail": scan,
            "keepLicense": scan or auto_send,
            "externalAccess": False,
            "autoSignature": auto_send,
            "mfaException": is_service,
        },
        "updatedAt": "2026-06-14T00:00:00.000Z",
    })

# Quelques membres / liens réalistes sur des boîtes partagées
by_email = {u["googleEmail"].lower(): u for u in users}
def add_members(email, members, right="accès complet"):
    u = by_email.get(email.lower())
    if u:
        u["mailbox"]["members"] = members
        u["mailbox"]["memberRight"] = right

add_members("contact@corthay.com", ["office@corthay.com", "patricia@corthay.com"])
add_members("corthay.order@corthay.com", ["office@corthay.com", "lisa@corthay.com"])
add_members("maison.volney@corthay.com", ["lisa@corthay.com", "assistantretail1@corthay.com"])
add_members("eshop@corthay.com", ["office@corthay.com"])

# --- Référentiel licences (prix réels onglet « Suivi Licences ») ------------
license_types = [
    {"code": "P1", "label": "Microsoft 365 Business Premium", "storageGB": 50,
     "price": {"EA_PA": 19.10, "EA_PM": 20.06, "EM_PM": 22.92}, "packBeCloud": 42.50, "countedInLicenses": True},
    {"code": "P2", "label": "Microsoft 365 F3 + Exchange Online Plan 1", "storageGB": 50,
     "price": {"EA_PA": 10.40, "EA_PM": 10.93, "EM_PM": 12.48}, "packBeCloud": 0, "countedInLicenses": True},
    {"code": "P3", "label": "Microsoft 365 F3", "storageGB": 2,
     "price": {"EA_PA": 6.90, "EA_PM": 7.25, "EM_PM": 8.28}, "packBeCloud": 0, "countedInLicenses": True},
    {"code": "P4a", "label": "Exchange Online Plan 1", "storageGB": 50,
     "price": {"EA_PA": 3.50, "EA_PM": 3.68, "EM_PM": 4.20}, "packBeCloud": 0, "countedInLicenses": True},
    {"code": "P4b", "label": "Exchange Online Kiosk", "storageGB": 2,
     "price": {"EA_PA": 1.73, "EA_PM": 1.82, "EM_PM": 2.08}, "packBeCloud": 0, "countedInLicenses": True},
    {"code": "SHARED", "label": "Boîte partagée (gratuite ≤ 50 Go)", "storageGB": 50,
     "price": {"EA_PA": 0, "EA_PM": 0, "EM_PM": 0}, "packBeCloud": 0, "countedInLicenses": False},
]

# --- Jalons projet (cf. §3.5) ------------------------------------------------
milestones = [
    {"id": "m2", "step": 2, "title": "Étape 2 — Cadrage", "date": "2026-06-11",
     "startTime": "15:30", "endTime": "16:30", "status": "fait", "dependencies": [], "blockingRisks": ""},
    {"id": "m3", "step": 3, "title": "Étape 3 — ATP", "date": "2026-06-12",
     "startTime": "17:00", "endTime": "17:30", "status": "fait", "dependencies": ["m2"], "blockingRisks": ""},
    {"id": "m4", "step": 4, "title": "Étape 4 — Préparation DNS / utilisateurs / groupes", "date": "2026-06-15",
     "startTime": "10:00", "endTime": "12:00", "status": "en cours", "dependencies": ["m3"],
     "blockingRisks": "Comptes à arbitrer non tranchés, MFA en retard."},
    {"id": "m5", "step": 5, "title": "Étape 5 — Basculement DNS", "date": "2026-06-25",
     "startTime": "17:00", "endTime": "17:30", "status": "à venir", "dependencies": ["m4"],
     "blockingRisks": "DNS non préparé, grosses boîtes non nettoyées."},
    {"id": "m6", "step": 6, "title": "Étape 6 — Reconnexion utilisateurs", "date": "2026-06-26",
     "startTime": "09:30", "endTime": "12:30", "status": "à venir", "dependencies": ["m5"],
     "blockingRisks": "MFA non configurée avant reconnexion."},
    {"id": "m7", "step": 7, "title": "Étape 7 — Validation", "date": "2026-06-30",
     "startTime": "14:00", "endTime": "15:00", "status": "à venir", "dependencies": ["m6"], "blockingRisks": ""},
]

# --- Tâches préchargées (cf. §3.4) ------------------------------------------
def task(i, title, cat, owner, prio, desc="", entity=None, status="à faire"):
    return {"id": f"t{i:03d}", "title": title, "description": desc, "category": cat,
            "owner": owner, "priority": prio, "status": status, "dueDate": None,
            "doneDate": None, "linkedEntity": entity, "comment": "", "reference": ""}

tasks = [
    task(1, "Compléter le fichier de migration", "utilisateur", "Julien", "haute"),
    task(2, "Finir les listes de distribution", "liste distribution", "Julien", "normale"),
    task(3, "Ajouter les licences manquantes", "licence", "Julien", "haute"),
    task(4, "Color-coder les boîtes partagées", "boîte partagée", "Julien", "basse"),
    task(5, "Demander le nettoyage à Xavier", "nettoyage", "Julien", "critique",
         "Boîte 131 Go à ramener sous 45 Go.", "xavier@corthay.com"),
    task(6, "Nettoyer les boîtes 68 Go / 108 Go sous 45 Go", "nettoyage", "utilisateur", "haute"),
    task(7, "Prévoir Exchange Plan 2 pour boîtes > 50 Go / proches 100 Go", "licence", "BeCloud", "haute"),
    task(8, "Configurer Google Workspace OSI / API", "DNS", "BeCloud", "haute"),
    task(9, "Préparer la bascule DNS", "DNS", "BeCloud", "critique"),
    task(10, "Prévenir les utilisateurs : ne plus travailler sur les boîtes après bascule", "communication", "Julien", "haute"),
    task(11, "Accompagner la reconnexion Outlook", "support", "BeCloud", "normale"),
    task(12, "Vérifier envoi / réception", "support", "utilisateur", "normale"),
    task(13, "Vérifier l'historique mail", "support", "utilisateur", "normale"),
    task(14, "Suivre les cas MFA bloqués", "MFA", "BeCloud", "haute"),
    task(15, "Ouvrir tickets support (Outlook / Excel / add-ins)", "support", "BeCloud", "normale"),
]

# --- Listes de distribution --------------------------------------------------
distribution_lists = [
    {"id": "dl1", "name": "Ventes quotidiennes boutiques", "address": "ventes-quotidiennes@corthay.com",
     "internalMembers": ["office@corthay.com", "xavier@corthay.com", "edouard@corthay.com",
                          "maison.london@corthay.com", "assistantretail1@corthay.com"],
     "externalMembers": [], "allowExternalSenders": False,
     "usage": "Distribution des chiffres de vente quotidiens — Paris / Hong Kong / Londres.",
     "creationStatus": "à créer", "remarks": "Demandée par la direction."},
]

settings = {
    "currency": "EUR",
    "googleWorkspacePerUser": 15.60,
    "googleWorkspaceUsers": 45,
    "dropboxMonthly": 9.99,
    "defaultEngagement": "annuel",
    "defaultPayment": "mensuel",
    "defaultAuthor": "Julien",
    "dnsProvider": "Gandi",
}

audit_log = [
    {"id": "a001", "date": "2026-06-14T08:00:00.000Z", "author": "Julien", "action": "import",
     "entity": "BECLOUD ANALYSE.xlsx", "field": "—", "oldValue": "", "newValue": "45 comptes importés"},
]

db = {
    "users": users,
    "licenseTypes": license_types,
    "tasks": tasks,
    "risks": [],
    "distributionLists": distribution_lists,
    "milestones": milestones,
    "auditLog": audit_log,
    "settings": settings,
}

os.makedirs(os.path.dirname(os.path.abspath(OUT)), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(db, f, ensure_ascii=False, indent=2)

# Petit résumé de contrôle
from collections import Counter
pc = Counter(u["licenseProfile"] for u in users)
print(f"OK — {len(users)} comptes écrits dans {OUT}")
print("Répartition profils:", dict(pc))
print("Boîtes >100 Go:", [u['googleEmail'] for u in users if u['mailboxSizeGB']>100])
print("Boîtes >50 Go:", [f"{u['googleEmail']}={u['mailboxSizeGB']}" for u in users if u['mailboxSizeGB']>50])
print("VIP:", [u['googleEmail'] for u in users if u['vip']])
